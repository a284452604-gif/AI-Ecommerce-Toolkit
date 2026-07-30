"""数据导出页面：将数据库中的分析/优化记录导出为 Excel 或 CSV

支持:
    1. 导出类型选择：分析记录 / 优化记录
    2. 日期范围筛选
    3. 平台/风格筛选
    4. 格式选择：Excel (.xlsx) / CSV (.csv)
    5. 数据预览表格
"""

from __future__ import annotations

import csv
import os
from datetime import datetime, timedelta

from PySide6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QLabel, QWidget, QScrollArea,
    QPushButton, QSplitter, QFrame,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QGroupBox, QGridLayout, QAbstractItemView,
    QMessageBox, QComboBox, QDateEdit, QCheckBox,
    QFileDialog,
)
from PySide6.QtCore import Qt, QDate, Signal
from PySide6.QtGui import QFont

from framework.base_page import BasePage
from framework.app_context import AppContext


class DataExportPage(BasePage):
    """数据导出页面

    布局:
        QSplitter (水平)
        ├── 左侧面板: 导出配置
        │   ├── 导出类型选择
        │   ├── 日期范围筛选
        │   ├── 平台/风格筛选
        │   ├── 格式选择
        │   └── 导出按钮
        └── 右侧面板: 预览表格
    """

    def __init__(self, page_key: str, title: str, icon: str = ""):
        super().__init__(page_key, title, icon)
        self._context = AppContext()
        self._logger = self._context.logger.get_logger("data_export")

        # UI 引用
        self._type_combo: QComboBox | None = None
        self._date_from: QDateEdit | None = None
        self._date_to: QDateEdit | None = None
        self._use_date_filter: QCheckBox | None = None
        self._filter_combo: QComboBox | None = None
        self._filter_label: QLabel | None = None
        self._format_combo: QComboBox | None = None
        self._export_btn: QPushButton | None = None
        self._preview_btn: QPushButton | None = None
        self._preview_table: QTableWidget | None = None
        self._status_label: QLabel | None = None
        self._count_label: QLabel | None = None

        # 预览数据缓存
        self._preview_data: list[dict] = []

    def _setup_ui(self):
        splitter = QSplitter(Qt.Orientation.Horizontal)

        left_panel = self._create_left_panel()
        splitter.addWidget(left_panel)

        right_panel = self._create_right_panel()
        splitter.addWidget(right_panel)

        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        splitter.setSizes([350, 700])

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.addWidget(splitter)

    # ===================== 左侧面板 =====================

    def _create_left_panel(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(0, 0, 8, 0)
        layout.setSpacing(12)

        # -- 导出类型 --
        type_group = QGroupBox("导出类型")
        type_layout = QVBoxLayout(type_group)

        self._type_combo = QComboBox()
        self._type_combo.addItem("商品链接分析记录", "analysis")
        self._type_combo.addItem("AI 标题优化记录", "optimization")
        self._type_combo.currentIndexChanged.connect(self._on_type_changed)
        type_layout.addWidget(self._type_combo)

        layout.addWidget(type_group)

        # -- 日期范围 --
        date_group = QGroupBox("日期范围")
        date_layout = QGridLayout(date_group)
        date_layout.setSpacing(8)

        self._use_date_filter = QCheckBox("启用日期筛选")
        self._use_date_filter.toggled.connect(self._on_date_filter_toggled)
        date_layout.addWidget(self._use_date_filter, 0, 0, 1, 2)

        date_layout.addWidget(QLabel("从:"), 1, 0)
        self._date_from = QDateEdit()
        self._date_from.setCalendarPopup(True)
        self._date_from.setDate(QDate.currentDate().addDays(-30))
        self._date_from.setEnabled(False)
        date_layout.addWidget(self._date_from, 1, 1)

        date_layout.addWidget(QLabel("到:"), 2, 0)
        self._date_to = QDateEdit()
        self._date_to.setCalendarPopup(True)
        self._date_to.setDate(QDate.currentDate())
        self._date_to.setEnabled(False)
        date_layout.addWidget(self._date_to, 2, 1)

        layout.addWidget(date_group)

        # -- 平台/风格筛选 --
        filter_group = QGroupBox("数据筛选")
        filter_layout = QGridLayout(filter_group)
        filter_layout.setSpacing(8)

        self._filter_label = QLabel("平台:")
        filter_layout.addWidget(self._filter_label, 0, 0)
        self._filter_combo = QComboBox()
        self._update_filter_options("analysis")
        filter_layout.addWidget(self._filter_combo, 0, 1)

        layout.addWidget(filter_group)

        # -- 导出格式 --
        format_group = QGroupBox("导出格式")
        format_layout = QVBoxLayout(format_group)

        self._format_combo = QComboBox()
        self._format_combo.addItem("Excel (.xlsx)", "xlsx")
        self._format_combo.addItem("CSV (.csv)", "csv")
        format_layout.addWidget(self._format_combo)

        layout.addWidget(format_group)

        # -- 按钮 --
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self._preview_btn = QPushButton("预览数据")
        self._preview_btn.setMinimumHeight(36)
        self._preview_btn.clicked.connect(self._on_preview)

        self._export_btn = QPushButton("导出文件")
        self._export_btn.setObjectName("PrimaryButton")
        self._export_btn.setMinimumHeight(36)
        self._export_btn.clicked.connect(self._on_export)

        btn_layout.addWidget(self._preview_btn)
        btn_layout.addWidget(self._export_btn)
        layout.addLayout(btn_layout)

        # -- 状态 --
        self._status_label = QLabel("")
        self._status_label.setWordWrap(True)
        self._status_label.setStyleSheet("color: #888; font-size: 9pt;")
        layout.addWidget(self._status_label)

        layout.addStretch()
        return panel

    # ===================== 右侧面板 =====================

    def _create_right_panel(self) -> QWidget:
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(8, 0, 0, 0)
        layout.setSpacing(12)

        title_label = QLabel("数据预览")
        title_label.setObjectName("PageTitle")
        layout.addWidget(title_label)

        self._count_label = QLabel("")
        self._count_label.setStyleSheet("color: #888; font-size: 9pt;")
        layout.addWidget(self._count_label)

        self._preview_table = QTableWidget(0, 1)
        self._preview_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._preview_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._preview_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._preview_table.setAlternatingRowColors(True)
        self._preview_table.verticalHeader().setVisible(False)
        layout.addWidget(self._preview_table, stretch=1)

        return panel

    # ===================== 信号处理 =====================

    def _on_type_changed(self, index: int):
        export_type = self._type_combo.currentData()
        self._update_filter_options(export_type)

    def _update_filter_options(self, export_type: str):
        """根据导出类型更新筛选选项"""
        self._filter_combo.clear()
        if export_type == "analysis":
            self._filter_label.setText("平台:")
            self._filter_combo.addItem("全部平台", "")
            self._filter_combo.addItem("淘宝", "淘宝")
            self._filter_combo.addItem("天猫", "天猫")
            self._filter_combo.addItem("京东", "京东")
            self._filter_combo.addItem("拼多多", "拼多多")
            self._filter_combo.addItem("抖音", "抖音")
        else:
            self._filter_label.setText("风格:")
            self._filter_combo.addItem("全部风格", "")
            self._filter_combo.addItem("搜索优化", "seo")
            self._filter_combo.addItem("促销转化", "promotion")
            self._filter_combo.addItem("品牌调性", "brand")

    def _on_date_filter_toggled(self, enabled: bool):
        self._date_from.setEnabled(enabled)
        self._date_to.setEnabled(enabled)

    def _get_filters(self) -> dict:
        """获取当前筛选条件"""
        export_type = self._type_combo.currentData()
        filters = {"type": export_type}

        if self._use_date_filter.isChecked():
            filters["date_from"] = self._date_from.date().toString("yyyy-MM-dd")
            filters["date_to"] = self._date_to.date().toString("yyyy-MM-dd")

        if export_type == "analysis":
            filters["platform"] = self._filter_combo.currentData() or ""
        else:
            filters["style"] = self._filter_combo.currentData() or ""

        return filters

    def _on_preview(self):
        """预览数据"""
        try:
            filters = self._get_filters()
            db = self._context.db

            if filters["type"] == "analysis":
                data = db.export_analysis_to_list(
                    platform=filters.get("platform", ""),
                    date_from=filters.get("date_from", ""),
                    date_to=filters.get("date_to", ""),
                )
                self._preview_data = data
                self._show_analysis_preview(data)
            else:
                data = db.export_optimization_to_list(
                    style=filters.get("style", ""),
                    date_from=filters.get("date_from", ""),
                    date_to=filters.get("date_to", ""),
                )
                self._preview_data = data
                self._show_optimization_preview(data)

            self._count_label.setText(f"共 {len(data)} 条记录")
        except Exception as e:
            self._status_label.setText(f"预览失败: {e}")
            self._logger.error(f"数据预览失败: {e}")

    def _show_analysis_preview(self, data: list[dict]):
        """预览分析记录"""
        columns = ["时间", "平台", "商品ID", "标题", "价格", "店铺"]
        self._preview_table.setColumnCount(len(columns))
        self._preview_table.setHorizontalHeaderLabels(columns)
        self._preview_table.setRowCount(len(data))

        for row, record in enumerate(data):
            self._preview_table.setItem(row, 0, QTableWidgetItem(record.get("created_at", "")))
            self._preview_table.setItem(row, 1, QTableWidgetItem(record.get("platform", "")))
            self._preview_table.setItem(row, 2, QTableWidgetItem(record.get("product_id", "")))
            self._preview_table.setItem(row, 3, QTableWidgetItem(record.get("title", "")))
            self._preview_table.setItem(row, 4, QTableWidgetItem(record.get("price", "")))
            self._preview_table.setItem(row, 5, QTableWidgetItem(record.get("shop_name", "")))

        header = self._preview_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)

    def _show_optimization_preview(self, data: list[dict]):
        """预览优化记录"""
        columns = ["时间", "风格", "原标题", "优化标题", "Token"]
        self._preview_table.setColumnCount(len(columns))
        self._preview_table.setHorizontalHeaderLabels(columns)
        self._preview_table.setRowCount(len(data))

        for row, record in enumerate(data):
            self._preview_table.setItem(row, 0, QTableWidgetItem(record.get("created_at", "")))
            self._preview_table.setItem(row, 1, QTableWidgetItem(record.get("style_name", "")))
            self._preview_table.setItem(row, 2, QTableWidgetItem(record.get("original_title", "")))
            self._preview_table.setItem(row, 3, QTableWidgetItem(record.get("optimized_title", "")))
            self._preview_table.setItem(row, 4, QTableWidgetItem(str(record.get("tokens_used", 0))))

        header = self._preview_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)

    def _on_export(self):
        """执行导出"""
        # 先获取预览数据
        if not self._preview_data:
            self._on_preview()

        if not self._preview_data:
            QMessageBox.warning(self, "提示", "没有可导出的数据")
            return

        export_type = self._type_combo.currentData()
        fmt = self._format_combo.currentData()

        # 默认文件名
        today = datetime.now().strftime("%Y%m%d")
        prefix = "分析记录" if export_type == "analysis" else "优化记录"

        if fmt == "xlsx":
            default_name = f"{prefix}_{today}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "导出 Excel", default_name,
                "Excel 文件 (*.xlsx);;所有文件 (*)"
            )
            if not file_path:
                return
            try:
                self._export_excel(file_path, export_type)
            except ImportError:
                QMessageBox.critical(self, "错误",
                    "导出 Excel 需要安装 openpyxl 库。\n请运行: pip install openpyxl")
                return
        else:
            default_name = f"{prefix}_{today}.csv"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "导出 CSV", default_name,
                "CSV 文件 (*.csv);;所有文件 (*)"
            )
            if not file_path:
                return
            self._export_csv(file_path, export_type)

        if os.path.exists(file_path):
            self._status_label.setText(
                f"✅ 导出成功！\n文件: {file_path}\n共 {len(self._preview_data)} 条记录"
            )
            self._logger.info(f"导出成功: {file_path} ({len(self._preview_data)} 条)")
        else:
            self._status_label.setText("❌ 导出失败: 文件未生成")
            self._logger.error("导出失败: 文件未生成")

    def _export_excel(self, file_path: str, export_type: str):
        """导出为 Excel 文件"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "分析记录" if export_type == "analysis" else "优化记录"

        # 表头样式
        header_font = Font(name="Microsoft YaHei UI", bold=True, size=11)
        header_fill = PatternFill(start_color="4A47A3", end_color="4A47A3", fill_type="solid")
        header_font_white = Font(name="Microsoft YaHei UI", bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        if export_type == "analysis":
            headers = ["时间", "平台", "商品ID", "标题", "价格", "店铺", "描述", "耗时(秒)", "链接"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for record in self._preview_data:
                ws.append([
                    record.get("created_at", ""),
                    record.get("platform", ""),
                    record.get("product_id", ""),
                    record.get("title", ""),
                    record.get("price", ""),
                    record.get("shop_name", ""),
                    record.get("description", ""),
                    record.get("fetch_time", 0),
                    record.get("url", ""),
                ])

            widths = [20, 10, 20, 50, 12, 20, 40, 10, 50]
        else:
            headers = ["时间", "风格", "原标题", "优化标题", "SEO关键词", "优化理由", "Token"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for record in self._preview_data:
                keywords = record.get("seo_keywords", [])
                kw_str = ", ".join(keywords) if keywords else ""
                ws.append([
                    record.get("created_at", ""),
                    record.get("style_name", ""),
                    record.get("original_title", ""),
                    record.get("optimized_title", ""),
                    kw_str,
                    record.get("improvement_reason", ""),
                    record.get("tokens_used", 0),
                ])

            widths = [20, 12, 50, 50, 30, 40, 10]

        # 设置列宽
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        # 数据行边框
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        wb.save(file_path)

    def _export_csv(self, file_path: str, export_type: str):
        """导出为 CSV 文件"""
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            if export_type == "analysis":
                writer = csv.writer(f)
                writer.writerow(["时间", "平台", "商品ID", "标题", "价格", "店铺", "描述", "耗时(秒)", "链接"])
                for record in self._preview_data:
                    writer.writerow([
                        record.get("created_at", ""),
                        record.get("platform", ""),
                        record.get("product_id", ""),
                        record.get("title", ""),
                        record.get("price", ""),
                        record.get("shop_name", ""),
                        record.get("description", ""),
                        record.get("fetch_time", 0),
                        record.get("url", ""),
                    ])
            else:
                writer = csv.writer(f)
                writer.writerow(["时间", "风格", "原标题", "优化标题", "SEO关键词", "优化理由", "Token"])
                for record in self._preview_data:
                    keywords = record.get("seo_keywords", [])
                    kw_str = ", ".join(keywords) if keywords else ""
                    writer.writerow([
                        record.get("created_at", ""),
                        record.get("style_name", ""),
                        record.get("original_title", ""),
                        record.get("optimized_title", ""),
                        kw_str,
                        record.get("improvement_reason", ""),
                        record.get("tokens_used", 0),
                    ])
